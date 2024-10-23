import openpyxl as xl
from openpyxl.chart import BarChart, Reference
#from the module chart in the package openpyxl we are importing two classes-BarChart & Reference 

wb=xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row+1):
      cell=sheet.cell(row,3)
      p = cell.value*0.9 
      c= sheet.cell(row,4)
      c.value=p  

value= Reference(sheet, min_row=2, max_row= sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(value)
sheet.add_chart(chart, 'e2')
wb.save('transactions2.xlsx') 

